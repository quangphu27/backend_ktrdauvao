"""API bài kiểm tra tự tạo — admin CRUD + học sinh làm bài."""

from datetime import datetime
from io import BytesIO

from flask import Blueprint, current_app, jsonify, request, send_file
from flask_jwt_extended import get_jwt, jwt_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database import col, parse_oid
from services.cloudinary_service import upload_image_file
from services.quiz_service import (
    attempt_to_dict,
    generate_quiz_slug,
    grade_attempt,
    quiz_to_dict,
    sanitize_questions,
)

quizzes_bp = Blueprint("quizzes", __name__, url_prefix="/api/quizzes")


def _admin_required():
    return get_jwt().get("role") == "admin"


def _unique_slug():
    for _ in range(20):
        slug = generate_quiz_slug()
        if not col("quizzes").find_one({"slug": slug}):
            return slug
    return generate_quiz_slug(12)


# ── Public ──────────────────────────────────────────────

@quizzes_bp.route("/by-slug/<slug>", methods=["GET"])
def get_quiz_public(slug):
    doc = col("quizzes").find_one({"slug": slug, "is_active": True})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra hoặc đã tắt"}), 404
    data = quiz_to_dict(doc, include_answers=False)
    # Không trả danh sách câu hỏi đầy đủ ở bước info — chỉ meta
    data["questions"] = []
    return jsonify(data)


@quizzes_bp.route("/by-slug/<slug>/start", methods=["POST"])
def start_quiz(slug):
    doc = col("quizzes").find_one({"slug": slug, "is_active": True})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra hoặc đã tắt"}), 404
    return jsonify(quiz_to_dict(doc, include_answers=False))


@quizzes_bp.route("/by-slug/<slug>/submit", methods=["POST"])
def submit_quiz(slug):
    doc = col("quizzes").find_one({"slug": slug, "is_active": True})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra hoặc đã tắt"}), 404

    data = request.get_json() or {}
    name = (data.get("student_name") or "").strip()
    grade = (data.get("student_grade") or data.get("grade") or "").strip()
    phone = (data.get("student_phone") or data.get("phone") or "").strip().replace(" ", "")
    answers = data.get("answers") or {}
    duration = int(data.get("duration_seconds") or 0)

    if not name or not grade:
        return jsonify({"message": "Vui lòng nhập họ tên và lớp"}), 400

    graded = grade_attempt(doc, answers)
    attempt = {
        "quiz_id": str(doc["_id"]),
        "quiz_title": doc.get("title"),
        "quiz_slug": doc.get("slug"),
        "student_name": name,
        "student_grade": grade,
        "student_phone": phone,
        "answers": {str(k): v for k, v in answers.items()},
        "details": graded["details"],
        "earned": graded["earned"],
        "max_score": graded["max_score"],
        "score": graded["score"],
        "duration_seconds": duration,
        "submitted_at": datetime.utcnow(),
    }
    result = col("quiz_attempts").insert_one(attempt)
    attempt["_id"] = result.inserted_id
    return jsonify({
        "message": "Nộp bài thành công!",
        "attempt": attempt_to_dict(attempt, include_details=True),
    }), 201


@quizzes_bp.route("/attempts/<attempt_id>", methods=["GET"])
def get_attempt_public(attempt_id):
    doc = col("quiz_attempts").find_one({"_id": parse_oid(attempt_id)})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài làm"}), 404
    return jsonify(attempt_to_dict(doc, include_details=True))


# ── Admin ───────────────────────────────────────────────

@quizzes_bp.route("/admin", methods=["GET"])
@jwt_required()
def admin_list_quizzes():
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    docs = list(col("quizzes").find().sort("created_at", -1))
    out = []
    for d in docs:
        item = quiz_to_dict(d, include_answers=False)
        item["questions"] = []
        item["attempt_count"] = col("quiz_attempts").count_documents({"quiz_id": str(d["_id"])})
        out.append(item)
    return jsonify(out)


@quizzes_bp.route("/admin", methods=["POST"])
@jwt_required()
def admin_create_quiz():
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    data = request.get_json() or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"message": "Vui lòng nhập tên bài kiểm tra"}), 400

    try:
        questions = sanitize_questions(data.get("questions") or [])
    except ValueError as e:
        return jsonify({"message": str(e)}), 400

    now = datetime.utcnow()
    doc = {
        "title": title,
        "description": (data.get("description") or "").strip(),
        "slug": _unique_slug(),
        "duration_minutes": max(0, int(data.get("duration_minutes") or 0)),
        "is_active": bool(data.get("is_active", True)),
        "questions": questions,
        "created_at": now,
        "updated_at": now,
    }
    result = col("quizzes").insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify(quiz_to_dict(doc, include_answers=True)), 201


@quizzes_bp.route("/admin/<quiz_id>", methods=["GET"])
@jwt_required()
def admin_get_quiz(quiz_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    doc = col("quizzes").find_one({"_id": parse_oid(quiz_id)})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra"}), 404
    return jsonify(quiz_to_dict(doc, include_answers=True))


@quizzes_bp.route("/admin/<quiz_id>", methods=["PUT"])
@jwt_required()
def admin_update_quiz(quiz_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    doc = col("quizzes").find_one({"_id": parse_oid(quiz_id)})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra"}), 404

    data = request.get_json() or {}
    updates = {"updated_at": datetime.utcnow()}

    if "title" in data:
        title = (data.get("title") or "").strip()
        if not title:
            return jsonify({"message": "Tên bài không được trống"}), 400
        updates["title"] = title
    if "description" in data:
        updates["description"] = (data.get("description") or "").strip()
    if "duration_minutes" in data:
        updates["duration_minutes"] = max(0, int(data.get("duration_minutes") or 0))
    if "is_active" in data:
        updates["is_active"] = bool(data.get("is_active"))
    if "questions" in data:
        try:
            updates["questions"] = sanitize_questions(data.get("questions") or [])
        except ValueError as e:
            return jsonify({"message": str(e)}), 400

    col("quizzes").update_one({"_id": doc["_id"]}, {"$set": updates})
    doc = col("quizzes").find_one({"_id": doc["_id"]})
    return jsonify(quiz_to_dict(doc, include_answers=True))


@quizzes_bp.route("/admin/<quiz_id>", methods=["DELETE"])
@jwt_required()
def admin_delete_quiz(quiz_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    oid = parse_oid(quiz_id)
    doc = col("quizzes").find_one({"_id": oid})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài kiểm tra"}), 404
    col("quiz_attempts").delete_many({"quiz_id": str(oid)})
    col("quizzes").delete_one({"_id": oid})
    return jsonify({"message": "Đã xóa bài kiểm tra và các bài làm liên quan"})


@quizzes_bp.route("/admin/upload-image", methods=["POST"])
@jwt_required()
def admin_upload_image():
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    f = request.files.get("file") or request.files.get("image")
    if not f or not f.filename:
        return jsonify({"message": "Chưa chọn ảnh"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in current_app.config.get("ALLOWED_EXTENSIONS", set()):
        return jsonify({"message": "Chỉ chấp nhận ảnh png/jpg/jpeg/gif/webp"}), 400
    try:
        meta = upload_image_file(f, folder="quiz_images")
    except Exception as exc:
        return jsonify({"message": f"Upload ảnh thất bại: {exc}"}), 500
    return jsonify(meta)


@quizzes_bp.route("/admin/<quiz_id>/attempts", methods=["GET"])
@jwt_required()
def admin_list_attempts(quiz_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    docs = col("quiz_attempts").find({"quiz_id": str(quiz_id)}).sort("submitted_at", -1)
    return jsonify([attempt_to_dict(d) for d in docs])


@quizzes_bp.route("/admin/attempts/<attempt_id>", methods=["GET"])
@jwt_required()
def admin_get_attempt(attempt_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    doc = col("quiz_attempts").find_one({"_id": parse_oid(attempt_id)})
    if not doc:
        return jsonify({"message": "Không tìm thấy bài làm"}), 404
    return jsonify(attempt_to_dict(doc, include_details=True))


@quizzes_bp.route("/admin/attempts/<attempt_id>", methods=["DELETE"])
@jwt_required()
def admin_delete_attempt(attempt_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    result = col("quiz_attempts").delete_one({"_id": parse_oid(attempt_id)})
    if result.deleted_count == 0:
        return jsonify({"message": "Không tìm thấy bài làm"}), 404
    return jsonify({"message": "Đã xóa bài làm"})


@quizzes_bp.route("/admin/<quiz_id>/export", methods=["GET"])
@jwt_required()
def admin_export_attempts(quiz_id):
    if not _admin_required():
        return jsonify({"message": "Không có quyền truy cập"}), 403
    quiz = col("quizzes").find_one({"_id": parse_oid(quiz_id)})
    if not quiz:
        return jsonify({"message": "Không tìm thấy bài kiểm tra"}), 404

    attempts = list(col("quiz_attempts").find({"quiz_id": str(quiz["_id"])}).sort("submitted_at", -1))
    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua"
    headers = ["STT", "Hoc sinh", "Lop", "SDT", "Diem %", "Dung/Tong", "Thoi gian (giay)", "Ngay nop"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4A90D9")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for i, a in enumerate(attempts, 1):
        submitted = a.get("submitted_at")
        date_str = submitted.strftime("%d/%m/%Y %H:%M") if hasattr(submitted, "strftime") else str(submitted or "")
        ws.append([
            i,
            a.get("student_name"),
            a.get("student_grade"),
            a.get("student_phone") or "",
            a.get("score"),
            f"{a.get('earned')}/{a.get('max_score')}",
            a.get("duration_seconds") or 0,
            date_str,
        ])

    ws2 = wb.create_sheet("Chi tiet")
    ws2.append(["Hoc sinh", "Lop", "Cau hoi", "Loai", "Tra loi", "Dap an dung", "Ket qua", "Diem"])
    for col_idx in range(1, 9):
        cell = ws2.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    for a in attempts:
        for d in a.get("details") or []:
            ws2.append([
                a.get("student_name"),
                a.get("student_grade"),
                d.get("content") or "(có ảnh)",
                d.get("type"),
                d.get("student_answer") or "",
                d.get("correct_answer") or "",
                "Dung" if d.get("is_correct") else "Sai",
                d.get("points") if d.get("is_correct") else 0,
            ])
            ws2.cell(ws2.max_row, 3).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 50

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in (quiz.get("title") or "quiz"))[:40]
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ket_qua_{safe_title}.xlsx",
    )
