import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _public_detail_url(test_id, frontend_url):
    base = (frontend_url or "https://kiemtradauvao.vercel.app").rstrip("/")
    return f"{base}/bai-lam/{test_id}"


def export_tests_excel(tests, frontend_url=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua kiem tra"
    headers = [
        "ID", "Học sinh", "Lớp", "Số điện thoại", "Khóa học", "Điểm", "Lộ trình",
        "Nhận xét", "Thời gian (giây)", "Ngày nộp", "Chi tiết bài làm",
    ]
    ws.append(headers)
    link_font = Font(color="0563C1", underline="single")
    detail_col = len(headers)

    for row_idx, t in enumerate(tests, start=2):
        submitted = t.get("submitted_at")
        url = t.get("detail_url") or _public_detail_url(t.get("id"), frontend_url)
        ws.append([
            t.get("id"),
            t.get("student_name"),
            t.get("student_grade"),
            t.get("student_phone"),
            t.get("course_name"),
            t.get("score"),
            t.get("recommendation"),
            t.get("comment"),
            t.get("duration_seconds"),
            submitted.strftime("%d/%m/%Y %H:%M") if submitted else "",
            url,
        ])
        cell = ws.cell(row=row_idx, column=detail_col, value=url)
        cell.hyperlink = url
        cell.font = link_font

    ws.column_dimensions[get_column_letter(detail_col)].width = 58
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _safe_filename(name):
    if not name:
        return "hoc_sinh"
    s = re.sub(r'[<>:"/\\|?*]', "", name.strip())
    s = re.sub(r"\s+", "_", s)
    return s[:60] or "hoc_sinh"


def _safe_phone(phone):
    digits = re.sub(r"\D", "", str(phone or ""))
    return digits or "0000000000"


def feedback_image_filename(test):
    """Đặt tên file: ten_lop_SDT.png"""
    name = _safe_filename(test.get("student_name"))
    grade = _safe_filename(test.get("student_grade") or "lop")
    phone = _safe_phone(test.get("student_phone"))
    return f"{name}_{grade}_{phone}.png"


def _clean_question_text(content):
    if not content:
        return ""
    return content.replace("`", "").strip()


LEVEL_VI = {"easy": "Dễ", "medium": "TB", "hard": "Khó"}
CATEGORY_VI = {
    "logic": "Logic",
    "observation": "Quan sát",
    "problem_solving": "Giải quyết VĐ",
    "sequencing": "Trình tự",
    "memory": "Ghi nhớ/Code",
    "math": "Tính toán",
    "algorithm": "Thuật toán",
}


def export_student_detail_excel(test, detailed_comment, frontend_url=None):
    """Xuất Excel chi tiết 1 học sinh: thông tin + nhận xét + từng câu."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bai lam"

    header_fill = PatternFill("solid", fgColor="4A90D9")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    info_rows = [
        ("Học sinh", test.get("student_name")),
        ("Lớp", test.get("student_grade")),
        ("Số điện thoại", test.get("student_phone")),
        ("Khóa học", test.get("course_name")),
        ("Điểm", test.get("score")),
        ("Lộ trình đề xuất", test.get("recommendation")),
        ("Ngày nộp", test.get("submitted_at", "")[:10] if test.get("submitted_at") else ""),
        ("Link chi tiết", test.get("detail_url") or _public_detail_url(test.get("id"), frontend_url)),
    ]
    for label, value in info_rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.append([])
    ws.append(["NHẬN XÉT CHI TIẾT"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append([detailed_comment])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
    ws.cell(row=ws.max_row, column=1).alignment = wrap

    ws.append([])
    headers = ["STT", "Mức độ", "Chủ đề", "Câu hỏi", "Trả lời", "Kết quả", "Đáp án đúng"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    details = test.get("details") or []
    for i, d in enumerate(details, 1):
        result = "Đúng" if d.get("is_correct") else ("Bỏ trống" if not d.get("answer_text") else "Sai")
        ws.append([
            i,
            LEVEL_VI.get(d.get("level"), d.get("level")),
            CATEGORY_VI.get(d.get("category"), d.get("category")),
            _clean_question_text(d.get("question_content")),
            d.get("answer_text") or "",
            result,
            d.get("correct_answer") or "",
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = wrap
        if not d.get("is_correct"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FEE2E2")

    widths = [6, 8, 14, 50, 28, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_student_detail_to_path(test, detailed_comment, out_dir, frontend_url=None):
    """Lưu file Excel ra thư mục, trả về đường dẫn file."""
    import os

    os.makedirs(out_dir, exist_ok=True)
    name = _safe_filename(test.get("student_name"))
    tid = (test.get("id") or "")[:8]
    path = os.path.join(out_dir, f"{name}_{tid}.xlsx")
    data = export_student_detail_excel(test, detailed_comment, frontend_url)
    with open(path, "wb") as f:
        f.write(data.read())
    return path


def export_all_students_workbook(tests_with_comments, frontend_url=None):
    """Xuất 1 file Excel: sheet Tổng hợp nhận xét + sheet Chi tiết từng câu."""
    wb = Workbook()
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill("solid", fgColor="4A90D9")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")

    # --- Sheet 1: Tổng hợp ---
    ws_sum = wb.active
    ws_sum.title = "Tong hop nhan xet"
    sum_headers = [
        "STT", "Học sinh", "Lớp", "SĐT", "Khóa học", "Điểm", "Đúng/Tổng",
        "Lộ trình", "Thời gian làm bài", "Ngày nộp", "Link bài làm", "Nhận xét chi tiết",
    ]
    ws_sum.append(sum_headers)
    for col in range(1, len(sum_headers) + 1):
        c = ws_sum.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(tests_with_comments, 1):
        test = item["test"]
        comment = item["comment"]
        details = test.get("details") or []
        correct = sum(1 for d in details if d.get("is_correct"))
        total = len(details)
        dur = test.get("duration_seconds") or 0
        m, s = divmod(int(dur), 60)
        dur_str = f"{m}p{s}s" if m else f"{s}s"
        submitted = (test.get("submitted_at") or "")[:10]
        url = test.get("detail_url") or _public_detail_url(test.get("id"), frontend_url)

        ws_sum.append([
            idx,
            test.get("student_name"),
            test.get("student_grade"),
            test.get("student_phone"),
            test.get("course_name"),
            test.get("score"),
            f"{correct}/{total}",
            test.get("recommendation"),
            dur_str,
            submitted,
            url,
            comment,
        ])
        row = ws_sum.max_row
        for col in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row, column=col).alignment = wrap
        link_cell = ws_sum.cell(row=row, column=11)
        link_cell.hyperlink = url
        link_cell.font = link_font
        ws_sum.row_dimensions[row].height = max(120, min(400, 15 + comment.count("\n") * 12))

    sum_widths = [5, 22, 10, 14, 10, 8, 10, 22, 12, 12, 42, 80]
    for i, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = "A2"

    # --- Sheet 2: Chi tiết từng câu ---
    ws_det = wb.create_sheet("Chi tiet tung cau")
    det_headers = [
        "Học sinh", "Lớp", "Khóa", "Điểm", "STT câu", "Mức độ", "Chủ đề",
        "Câu hỏi", "Trả lời", "Kết quả", "Đáp án đúng",
    ]
    ws_det.append(det_headers)
    for col in range(1, len(det_headers) + 1):
        c = ws_det.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for item in tests_with_comments:
        test = item["test"]
        name = test.get("student_name")
        grade = test.get("student_grade")
        course = test.get("course_name")
        score = test.get("score")
        for i, d in enumerate(test.get("details") or [], 1):
            result = "Đúng" if d.get("is_correct") else ("Bỏ trống" if not d.get("answer_text") else "Sai")
            ws_det.append([
                name, grade, course, score, i,
                LEVEL_VI.get(d.get("level"), d.get("level")),
                CATEGORY_VI.get(d.get("category"), d.get("category")),
                _clean_question_text(d.get("question_content")),
                d.get("answer_text") or "",
                result,
                d.get("correct_answer") or "",
            ])
            row = ws_det.max_row
            for col in range(1, len(det_headers) + 1):
                ws_det.cell(row=row, column=col).alignment = wrap
            if not d.get("is_correct"):
                for col in range(1, len(det_headers) + 1):
                    ws_det.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FEE2E2")

    det_widths = [20, 8, 8, 6, 6, 6, 14, 48, 24, 8, 24]
    for i, w in enumerate(det_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w
    ws_det.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_all_students_to_path(tests_with_comments, out_path, frontend_url=None):
    import os

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    data = export_all_students_workbook(tests_with_comments, frontend_url)
    with open(out_path, "wb") as f:
        f.write(data.read())
    return out_path


def export_tests_pdf(tests):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, spaceAfter=20)
    elements = [
        Paragraph("Bao cao ket qua kiem tra nang luc dau vao", title_style),
        Paragraph(f"Ngay xuat: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]
    data = [["ID", "Hoc sinh", "Lop", "SDT", "Khoa hoc", "Diem", "Lo trinh", "Ngay nop"]]
    for t in tests:
        submitted = t.get("submitted_at")
        data.append([
            str(t.get("id")),
            t.get("student_name"),
            t.get("student_grade"),
            t.get("student_phone"),
            t.get("course_name"),
            f"{t.get('score')}",
            t.get("recommendation"),
            submitted.strftime("%d/%m/%Y") if submitted else "",
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A90D9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F0F8FF")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F4FD")]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output
